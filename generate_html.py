"""Generate index.html from outages.xlsx for Gitee Pages hosting."""
import openpyxl
from datetime import datetime, timedelta

PYTHON = '/c/Users/Administrator/AppData/Local/Programs/Python/Python314/python.exe'

def serial_to_str(val):
    """Convert Excel serial/date to Chinese date string."""
    if val is None:
        return ''
    if isinstance(val, (int, float)) and val > 40000:
        d = datetime(1899, 12, 30) + timedelta(days=val)
        return f'{d.year}年{d.month}月{d.day}日'
    if isinstance(val, datetime):
        return f'{val.year}年{val.month}月{val.day}日'
    return str(val)

def safe_str(val, max_len=None):
    """Safe string conversion with optional truncation."""
    if val is None:
        return ''
    s = str(val).strip()
    if max_len and len(s) > max_len:
        return s[:max_len] + '…'
    return s

def read_sheet(ws, col_indices):
    """Read sheet data; returns list of dicts."""
    rows = []
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for key, col in col_indices.items():
            row_data[key] = ws.cell(r, col).value
        # Skip empty rows
        if row_data.get('date') is None:
            continue
        rows.append(row_data)
    return rows

def build_table_rows(rows, is_international=False):
    """Build HTML <tr> strings from row data."""
    html_rows = []
    for i, r in enumerate(rows):
        seq = safe_str(r.get('seq', ''))
        date = serial_to_str(r.get('date', ''))
        location = safe_str(r.get('location', ''))
        company = safe_str(r.get('company', ''))
        cause = safe_str(r.get('cause', ''))
        init_sys = safe_str(r.get('init_sys', ''))
        desc = safe_str(r.get('desc', ''), 80)
        impact = safe_str(r.get('impact', ''), 60)
        duration = safe_str(r.get('duration', ''))
        level = safe_str(r.get('level', ''))

        # Truncate long text for table display
        desc_display = safe_str(r.get('desc', ''), 80)
        impact_display = safe_str(r.get('impact', ''), 50)

        # Build tooltip with full text
        desc_full = safe_str(r.get('desc', '')).replace('"', '&quot;')
        impact_full = safe_str(r.get('impact', '')).replace('"', '&quot;')

        html_rows.append(f'''<tr>
            <td class="col-seq">{seq}</td>
            <td class="col-date">{date}</td>
            <td class="col-loc">{location}</td>
            <td class="col-company">{company}</td>
            <td class="col-cause">{cause}</td>
            <td class="col-sys">{init_sys}</td>
            <td class="col-desc" title="{desc_full}">{desc_display}</td>
            <td class="col-impact" title="{impact_full}">{impact_display}</td>
            <td class="col-dur">{duration}</td>
            <td class="col-level">{level}</td>
        </tr>''')
    return '\n'.join(html_rows)

def main():
    wb = openpyxl.load_workbook('outages.xlsx')
    neidi = wb.worksheets[0]
    guoji = wb.worksheets[1]

    # Column indices for reading
    neidi_cols = {
        'seq': 1, 'date': 2, 'location': 3, 'company': 4,
        'cause': 6, 'init_sys': 7, 'desc': 10, 'impact': 11,
        'duration': 13, 'level': 15,
    }
    guoji_cols = {
        'seq': 1, 'date': 2, 'location': 3, 'company': 4,
        'cause': 6, 'init_sys': 7, 'desc': 10, 'impact': 11,
        'duration': 13, 'level': 15,
    }

    neidi_rows = read_sheet(neidi, neidi_cols)
    guoji_rows = read_sheet(guoji, guoji_cols)

    # Reverse: newest first
    neidi_rows.reverse()
    guoji_rows.reverse()

    neidi_tbody = build_table_rows(neidi_rows, is_international=False)
    guoji_tbody = build_table_rows(guoji_rows, is_international=True)

    now_str = datetime.now().strftime('%Y年%m月%d日 %H:%M')

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>信息通信网络事故数据库</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", "宋体", sans-serif;
    background: #f0f2f5;
    color: #333;
    min-height: 100vh;
}}

/* Password Gate */
#pw-overlay {{
    position: fixed; top: 0; left: 0; width: 100%; height: 100%;
    background: linear-gradient(135deg, #1a3a5c 0%, #2d6aa0 100%);
    display: flex; align-items: center; justify-content: center;
    z-index: 9999;
}}
#pw-box {{
    background: white; border-radius: 12px; padding: 36px 32px;
    text-align: center; box-shadow: 0 8px 32px rgba(0,0,0,0.2);
    max-width: 360px; width: 90%;
}}
#pw-box h2 {{ color: #1a3a5c; margin-bottom: 8px; font-size: 1.3em; }}
#pw-box p {{ color: #888; font-size: 0.85em; margin-bottom: 20px; }}
#pw-box input {{
    width: 100%; padding: 10px 14px; border: 1px solid #d0d5dd;
    border-radius: 6px; font-size: 1em; text-align: center; outline: none;
    margin-bottom: 12px;
}}
#pw-box input:focus {{ border-color: #2d6aa0; box-shadow: 0 0 0 3px rgba(45,106,160,0.1); }}
#pw-box button {{
    width: 100%; padding: 10px; background: #2d6aa0; color: white;
    border: none; border-radius: 6px; font-size: 1em; cursor: pointer;
    font-weight: 600;
}}
#pw-box button:hover {{ background: #1a3a5c; }}
#pw-error {{ color: #c0392b; font-size: 0.85em; margin-top: 8px; display: none; }}

/* Main content (hidden until unlocked) */
#main-content {{ display: none; }}

.header {{
    background: linear-gradient(135deg, #1a3a5c 0%, #2d6aa0 100%);
    color: white;
    padding: 24px 20px;
    text-align: center;
}}
.header h1 {{ font-size: 1.6em; font-weight: 600; margin-bottom: 6px; }}
.header p {{ font-size: 0.85em; opacity: 0.8; }}

/* Tabs */
.tabs {{
    display: flex;
    justify-content: center;
    gap: 4px;
    padding: 16px 12px 0;
    max-width: 1300px;
    margin: 0 auto;
}}
.tab-btn {{
    padding: 10px 32px;
    border: none;
    background: #dce3ea;
    color: #555;
    font-size: 0.95em;
    cursor: pointer;
    border-radius: 8px 8px 0 0;
    font-weight: 500;
    transition: all 0.2s;
}}
.tab-btn.active {{
    background: white;
    color: #1a3a5c;
    font-weight: 700;
    box-shadow: 0 -2px 6px rgba(0,0,0,0.06);
}}
.tab-btn:hover {{ background: #e8ecf1; }}
.tab-btn .count {{
    font-size: 0.75em;
    color: #888;
    margin-left: 4px;
}}

/* Container */
.container {{
    max-width: 1300px;
    margin: 0 auto;
    padding: 12px;
}}
.table-wrap {{
    background: white;
    border-radius: 0 8px 8px 8px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.06);
    overflow: hidden;
}}

/* Search bar */
.search-bar {{
    padding: 14px 18px;
    border-bottom: 1px solid #eee;
    display: flex;
    gap: 10px;
    flex-wrap: wrap;
    align-items: center;
}}
.search-bar input {{
    flex: 1;
    min-width: 200px;
    padding: 8px 14px;
    border: 1px solid #d0d5dd;
    border-radius: 6px;
    font-size: 0.9em;
    outline: none;
}}
.search-bar input:focus {{ border-color: #2d6aa0; box-shadow: 0 0 0 3px rgba(45,106,160,0.1); }}
.search-bar .hint {{
    font-size: 0.78em;
    color: #999;
}}

/* Table */
table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.85em;
}}
thead th {{
    background: #f5f7fa;
    color: #444;
    font-weight: 600;
    padding: 10px 8px;
    text-align: left;
    border-bottom: 2px solid #e0e4e8;
    white-space: nowrap;
    cursor: pointer;
    user-select: none;
    position: sticky;
    top: 0;
}}
thead th:hover {{ background: #eef1f5; }}
thead th .arrow {{ font-size: 0.7em; margin-left: 3px; color: #bbb; }}
tbody td {{
    padding: 9px 8px;
    border-bottom: 1px solid #f0f0f0;
    vertical-align: top;
    line-height: 1.5;
}}
tbody tr:hover {{ background: #f8fafc; }}
.col-seq {{ width: 40px; text-align: center; color: #999; }}
.col-date {{ width: 90px; white-space: nowrap; font-size: 0.82em; }}
.col-loc {{ width: 70px; font-weight: 500; }}
.col-company {{ width: 100px; font-size: 0.82em; }}
.col-cause {{ width: 80px; font-size: 0.8em; }}
.col-sys {{ width: 80px; font-size: 0.8em; }}
.col-desc {{ min-width: 180px; }}
.col-impact {{ min-width: 130px; font-size: 0.82em; }}
.col-dur {{ width: 65px; text-align: right; font-size: 0.82em; white-space: nowrap; }}
.col-level {{ width: 50px; text-align: center; }}

.table-scroll {{ overflow-x: auto; max-height: 75vh; overflow-y: auto; }}

/* Footer */
.footer {{
    text-align: center;
    padding: 20px;
    color: #999;
    font-size: 0.78em;
}}
.footer a {{ color: #2d6aa0; text-decoration: none; }}

/* Mobile */
@media (max-width: 768px) {{
    .header h1 {{ font-size: 1.2em; }}
    .tab-btn {{ padding: 8px 16px; font-size: 0.85em; }}
    .col-cause, .col-sys, .col-level {{ display: none; }}
    .col-impact {{ display: none; }}
    .col-dur {{ font-size: 0.75em; }}
    table {{ font-size: 0.78em; }}
    thead th, tbody td {{ padding: 6px 5px; }}
    .col-company {{ max-width: 70px; font-size: 0.75em; }}
}}

/* Severity badge */
.badge-red {{ display: inline-block; padding: 2px 8px; border-radius: 10px; background: #fee2e2; color: #b91c1c; font-size: 0.8em; font-weight: 600; }}
.badge-yellow {{ display: inline-block; padding: 2px 8px; border-radius: 10px; background: #fef3c7; color: #92400e; font-size: 0.8em; font-weight: 600; }}
.badge-blue {{ display: inline-block; padding: 2px 8px; border-radius: 10px; background: #dbeafe; color: #1e40af; font-size: 0.8em; font-weight: 600; }}
</style>
</head>
<body>

<!-- Password Gate -->
<div id="pw-overlay">
<div id="pw-box">
    <h2>🔒 访问受限</h2>
    <p>请输入访问密码</p>
    <input type="password" id="pw-input" placeholder="密码" onkeydown="if(event.key==='Enter')checkPw()">
    <button onclick="checkPw()">确 定</button>
    <div id="pw-error">密码错误，请重试</div>
</div>
</div>

<!-- Main Content -->
<div id="main-content">

<div class="header">
    <h1>📡 信息通信网络事故数据库</h1>
    <p>全球通信网络重大中断事故记录 | 更新于 {now_str}</p>
</div>

<div class="tabs">
    <button class="tab-btn active" onclick="switchTab('neidi')">🇨🇳 内地事故 <span class="count">({len(neidi_rows)})</span></button>
    <button class="tab-btn" onclick="switchTab('guoji')">🌍 国际事故 <span class="count">({len(guoji_rows)})</span></button>
</div>

<div class="container">
<div class="table-wrap">

<div class="search-bar">
    <input type="text" id="searchBox" placeholder="搜索公司、地点、原因、关键词…" oninput="filterTable()">
    <span class="hint">共 <strong id="visibleCount">0</strong> 条记录</span>
</div>

<div class="table-scroll">

<!-- 内地表 -->
<table id="table-neidi">
<thead>
<tr>
    <th class="col-seq" onclick="sortTable(0,'neidi')"># <span class="arrow">⇅</span></th>
    <th class="col-date" onclick="sortTable(1,'neidi')">日期 <span class="arrow">⇅</span></th>
    <th class="col-loc">地点</th>
    <th class="col-company">公司</th>
    <th class="col-cause">原因</th>
    <th class="col-sys">初始故障系统</th>
    <th class="col-desc">基本情况</th>
    <th class="col-impact">影响</th>
    <th class="col-dur" onclick="sortTable(8,'neidi')">历时(分) <span class="arrow">⇅</span></th>
    <th class="col-level">等级</th>
</tr>
</thead>
<tbody>
{neidi_tbody}
</tbody>
</table>

<!-- 国际表 -->
<table id="table-guoji" style="display:none;">
<thead>
<tr>
    <th class="col-seq" onclick="sortTable(0,'guoji')"># <span class="arrow">⇅</span></th>
    <th class="col-date" onclick="sortTable(1,'guoji')">日期 <span class="arrow">⇅</span></th>
    <th class="col-loc">国家</th>
    <th class="col-company">公司</th>
    <th class="col-cause">原因</th>
    <th class="col-sys">初始故障系统</th>
    <th class="col-desc">基本情况</th>
    <th class="col-impact">影响</th>
    <th class="col-dur" onclick="sortTable(8,'guoji')">历时(分) <span class="arrow">⇅</span></th>
    <th class="col-level">等级</th>
</tr>
</thead>
<tbody>
{guoji_tbody}
</tbody>
</table>

</div>
</div>
</div>

<div class="footer">
    最后更新：{now_str}（北京时间） ·
    <a href="outages.xlsx" download>📥 下载完整 Excel</a> ·
    由 Claude Code 自动维护
</div>

</div><!-- #main-content -->

<script>
// Password check
const CORRECT_HASH = 'aa1cd1c664a86c8b1e7a21440e4817c94ce8a73c9f50e9be10c16d4fa16f1c36';

async function sha256(msg) {{
    const buf = new TextEncoder().encode(msg);
    const hash = await crypto.subtle.digest('SHA-256', buf);
    return Array.from(new Uint8Array(hash)).map(b => b.toString(16).padStart(2,'0')).join('');
}}

async function checkPw() {{
    const input = document.getElementById('pw-input').value;
    const hash = await sha256(input);
    if (hash === CORRECT_HASH) {{
        document.getElementById('pw-overlay').style.display = 'none';
        document.getElementById('main-content').style.display = '';
    }} else {{
        document.getElementById('pw-error').style.display = '';
        document.getElementById('pw-input').value = '';
        document.getElementById('pw-input').focus();
    }}
}}

document.getElementById('pw-input').focus();

<script>
// Tab switching
function switchTab(tab) {{
    document.getElementById('table-neidi').style.display = tab === 'neidi' ? '' : 'none';
    document.getElementById('table-guoji').style.display = tab === 'guoji' ? '' : 'none';
    document.querySelectorAll('.tab-btn').forEach((b, i) => {{
        b.classList.toggle('active', (tab === 'neidi' && i === 0) || (tab === 'guoji' && i === 1));
    }});
    document.getElementById('searchBox').value = '';
    updateCount(tab);
}}

// Filter table by search
function filterTable() {{
    const query = document.getElementById('searchBox').value.toLowerCase();
    const activeTab = document.getElementById('table-neidi').style.display === 'none' ? 'guoji' : 'neidi';
    const table = document.getElementById('table-' + activeTab);
    const rows = table.querySelectorAll('tbody tr');
    let visible = 0;
    rows.forEach(row => {{
        const text = row.textContent.toLowerCase();
        const match = !query || text.includes(query);
        row.style.display = match ? '' : 'none';
        if (match) visible++;
    }});
    document.getElementById('visibleCount').textContent = visible;
}}

// Sort table
function sortTable(colIdx, tab) {{
    const table = document.getElementById('table-' + tab);
    const tbody = table.querySelector('tbody');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    const isAsc = table.dataset.sortCol == colIdx && table.dataset.sortDir === 'asc';
    table.dataset.sortCol = colIdx;
    table.dataset.sortDir = isAsc ? 'desc' : 'asc';
    rows.sort((a, b) => {{
        let va = a.cells[colIdx].textContent.trim();
        let vb = b.cells[colIdx].textContent.trim();
        let na = parseFloat(va.replace(/[^0-9.-]/g, ''));
        let nb = parseFloat(vb.replace(/[^0-9.-]/g, ''));
        if (!isNaN(na) && !isNaN(nb)) return isAsc ? nb - na : na - nb;
        return isAsc ? vb.localeCompare(va, 'zh') : va.localeCompare(vb, 'zh');
    }});
    rows.forEach(r => tbody.appendChild(r));
}}

// Init
function updateCount(tab) {{
    const table = document.getElementById('table-' + tab);
    const rows = table.querySelectorAll('tbody tr');
    document.getElementById('visibleCount').textContent = rows.length;
}}
updateCount('neidi');
</script>

</body>
</html>'''

    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(html)

    print(f'index.html generated: {len(neidi_rows)} neidi + {len(guoji_rows)} guoji records')

if __name__ == '__main__':
    main()

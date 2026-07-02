import shutil, os
from datetime import datetime
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment

# Backup
backup_name = 'outages_backup_20260702.xlsx'
suffix = 'a'
while os.path.exists(backup_name):
    backup_name = f'outages_backup_20260702{suffix}.xlsx'
    suffix = chr(ord(suffix) + 1)
shutil.copy('outages.xlsx', backup_name)
print(f'Backup: {backup_name}')

wb = openpyxl.load_workbook('outages.xlsx')
guoji = wb.worksheets[1]
suoyin = wb.worksheets[2]

# Read index
idx_E, idx_F, idx_G, idx_H = set(), set(), set(), set()
for r in range(3, suoyin.max_row + 1):
    for val, s in [(suoyin.cell(r,1).value, idx_E), (suoyin.cell(r,2).value, idx_F),
                   (suoyin.cell(r,3).value, idx_G), (suoyin.cell(r,4).value, idx_H)]:
        if val and str(val).strip(): s.add(str(val).strip())

# Find last data row
last_data_row = guoji.max_row
while guoji.cell(last_data_row, 1).value is None and last_data_row > 0:
    last_data_row -= 1
seq = int(guoji.cell(last_data_row, 1).value or 0) + 1
row = last_data_row + 1
print(f'Starting row={row}, seq={seq}')

# Styles
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
song_font = Font(name='宋体', size=10)
cell_align = Alignment(wrap_text=True, vertical='center')

# New records
records = [
    {
        'A': seq,
        'B': '2026年6月29日',
        'C': '法国',
        'D': 'Orange（波及Sosh、YouPrice等MVNO）',
        'E': '通信企业',
        'F': '软件故障',
        'G': '其他',
        'H': '未及时发现',
        'I': '核心网技术故障',
        'J': '6月29日清晨起，Orange法国全国4G/5G移动网络中断约5-7小时，数百万用户无法上网和通话，国际漫游亦受影响。Orange确认系技术故障非网络攻击。',
        'K': '全国数百万用户移动网络中断，Downdetector峰值1600+报告，Sosh/YouPrice等MVNO同步瘫痪。',
        'L': '约5-7小时（清晨至11:00恢复）',
        'M': 360,
        'P': '核心网技术故障，具体原因Orange未公开披露',
        'Q': 'MVNO无独立回退能力，故障同步扩散',
        'T': 'Orange确认非网络攻击；用户可联系客服申请善意补偿',
        'U': '【待更新】根因未公开披露，法国电信监管机构Arcep或要求提交事故报告',
        'V': 'https://www.connexionfrance.com/news/what-is-known-about-orange-service-outages-this-week-in-france/800171',
        'W': 'https://www.tomsguide.fr/panne-orange-et-sosh-de-grosses-perturbations-signalees-ce-lundi/',
    },
    {
        'A': seq + 1,
        'B': '2026年7月1日',
        'C': '法国',
        'D': 'Orange（波及Sosh）',
        'E': '通信企业',
        'F': '软件故障',
        'G': '其他',
        'H': '未及时发现',
        'I': '移动通话故障',
        'J': '7月1日上午8:15至9:30，Orange法国再次发生全国移动通话中断约75分钟。Orange称与6月29日故障性质不同。三天内两次全国故障引发用户强烈不满。',
        'K': '全国数百万用户约75分钟无法拨打电话，Downdetector爆量投诉。',
        'L': '约75分钟（8:15-9:30）',
        'M': 75,
        'P': '移动语音核心网故障，具体原因未披露',
        'Q': '三天内第二次重大故障，网络可靠性受质疑',
        'T': 'Orange称两次故障性质不同，均非网络攻击',
        'U': '【待更新】根因未披露；三天内两次故障，Arcep或介入调查',
        'V': 'https://www.numerama.com/tech/2288757-panne-orange-les-signalements-explosent-ce-mercredi-1er-juillet.html',
        'W': 'https://www.sudouest.fr/economie/economie-du-numerique/nouvelle-panne-chez-orange-et-sosh-des-internautes-encore-prives-d-acces-au-reseau-et-d-appels-ce-mercredi-29741068.php',
    },
    {
        'A': seq + 2,
        'B': '2026年7月2日',
        'C': '乌克兰',
        'D': 'Undernet/Faust/IBnet/X-com/Vodafone Ukraine',
        'E': '通信企业',
        'F': '外部破坏',
        'G': '其他',
        'H': '设备断电',
        'I': '战争/外部破坏',
        'J': '7月1日夜间俄罗斯大规模导弹袭击基辅，一枚导弹击中关键数据中心，Undernet等多家ISP核心设备受损，骨干通信链路中断。基辅两岸多个城区固定互联网瘫痪。',
        'K': '基辅及周边地区固定互联网大面积中断，Undernet/Faust/IBnet/X-com等ISP用户断网，NAZK政府网站无法访问。至少13死90伤。',
        'L': '恢复时间未定（截至7月2日仍在抢修）',
        'P': '俄罗斯大规模导弹袭击击中民用数据中心',
        'Q': '骨干通信链路和ISP核心设备物理损坏，多运营商依赖同一数据中心',
        'T': '工程团队连夜抢修，移动网络依靠基站备电维持运转',
        'U': '【待更新】恢复时间未定，截至报道时仍在抢修中；Vodafone Ukraine家庭宽带和客服系统亦受影响',
        'V': 'https://en.interfax.com.ua/news/telecom/1181547.html',
        'W': 'https://24tv.ua/tech/vodafone-ukrayina-zbiy-merezhi-problemi-internetom-pislya-obstrilu_n3098351',
    },
    {
        'A': seq + 3,
        'B': '2026年7月1日',
        'C': '英国',
        'D': 'BT/Sky/EE/Vodafone/TalkTalk/Virgin',
        'E': '通信企业',
        'F': '设备设施故障',
        'G': '承载数通',
        'H': '冗余保护失败',
        'I': '交换局故障',
        'J': '7月1日起Carlisle地区Openreach交换局级网络故障，多运营商宽带和电话服务大面积中断。商家被迫改用现金交易，部分用户被告知次日恢复。',
        'K': 'Carlisle及周边地区多运营商宽带/电话中断，约10万人口城市受影响，商家被迫现金交易。',
        'L': '约24小时（部分用户次日恢复）',
        'M': 1440,
        'P': 'Openreach交换局级网络设备故障，影响所有租用其基础设施的运营商',
        'Q': '单一交换局无冗余备份，故障扩散至所有接入运营商',
        'T': 'EE工程师已展开排查，用户被建议避免反复重启设备',
        'U': '【待更新】Openreach未公布具体故障原因',
        'V': 'https://www.newsandstar.co.uk/news/26247822.carlisle-bt-sky-ee-talktalk-virgin-disruption/',
    },
]

col_map = {'A':1,'B':2,'C':3,'D':4,'E':5,'F':6,'G':7,'H':8,'I':9,'J':10,'K':11,'L':12,'M':13,'N':14,'O':15,'P':16,'Q':17,'R':18,'S':19,'T':20,'U':21,'V':22,'W':23}

for i, rec in enumerate(records):
    r = row + i
    for field, idx_set, name in [('E', idx_E, '性质'), ('F', idx_F, '原因'), ('G', idx_G, '初始故障系统'), ('H', idx_H, '事故扩大原因')]:
        val = rec.get(field)
        if val and val not in idx_set:
            print(f'  WARNING Row {r}: {name} "{val}" not in index!')

    for col_key, col_idx in col_map.items():
        val = rec.get(col_key)
        if val is not None:
            guoji.cell(r, col_idx, val)

    for c in range(1, 24):
        cell = guoji.cell(r, c)
        cell.border = thin
        cell.font = song_font
        cell.alignment = cell_align

    print(f'Row {r}: {rec["D"][:50]} written')

# Fix headers
if guoji.cell(1, 22).value is None or str(guoji.cell(1, 22).value).strip() == '':
    guoji.cell(1, 22).value = '消息来源'
if guoji.cell(1, 23).value is None or str(guoji.cell(1, 23).value).strip() == '':
    guoji.cell(1, 23).value = '原链接'

# Update Vodafone Row 67
guoji.cell(67, 21).value = '【待更新】ACMA新规6月30日已生效，Vodafone须强制公布事故详情；备份电源失效原因仍未披露'
guoji.cell(67, 21).font = song_font
guoji.cell(67, 21).alignment = cell_align
print('Vodafone Row 67 U updated')

wb.save('outages.xlsx')

# Verify
wb2 = openpyxl.load_workbook('outages.xlsx')
g2 = wb2.worksheets[1]
print('\n=== Verification ===')
for r in range(row, row + 4):
    print(f'Row {r}: B={g2.cell(r,2).value} | C={g2.cell(r,3).value} | D={g2.cell(r,4).value}')
    print(f'  E={g2.cell(r,5).value} F={g2.cell(r,6).value} G={g2.cell(r,7).value} H={g2.cell(r,8).value}')
    print(f'  V={g2.cell(r,22).value}')

print(f'\nTotal international rows: {g2.max_row}')
print('DONE')
